"""Enhanced FileOps — Full DC-parity file I/O.

Covers: read, write, read_multiple, list_dir, search, move, copy,
file_info, mkdir.  Adds Excel-sheet awareness to file_info,
read_multiple_files for batch reads, and image rendering support.
"""

import json
import shutil
import re
from pathlib import Path
from sassymcp.modules._security import validate_path


def _check_path(path: str) -> str | None:
    """Validate path against allowedDirectories. Returns error string or None."""
    ok, err = validate_path(path)
    if not ok:
        return err
    return None


def register(server):

    @server.tool()
    async def sassy_read_file(path: str, offset: int = 0, length: int = 1000) -> str:
        """Read file contents with line-based pagination.

        offset >= 0 : start from that line (0-based)
        offset < 0  : read last N lines (tail)
        length      : max lines to return (ignored when offset < 0)
        """
        err = _check_path(path)
        if err:
            return f"Error: {err}"
        p = Path(path)
        if not p.exists():
            return f"Error: {path} does not exist"
        if p.is_dir():
            return f"Error: {path} is a directory — use sassy_list_dir"
        try:
            text = p.read_text(encoding="utf-8", errors="replace")
        except (OSError, PermissionError) as e:
            return f"Error reading {path}: {e}"

        lines = text.splitlines()
        total = len(lines)

        if offset < 0:
            selected = lines[offset:]
            start_line = max(0, total + offset)
        else:
            selected = lines[offset : offset + length]
            start_line = offset

        remaining = total - start_line - len(selected)
        header = f"[Reading {len(selected)} lines from line {start_line} (total: {total}, {remaining} remaining)]"
        numbered = [f"  {start_line + i + 1:>5}\t{line}" for i, line in enumerate(selected)]
        return header + "\n" + "\n".join(numbered)

    @server.tool()
    async def sassy_read_multiple(paths: str) -> str:
        """Read multiple files at once. paths = JSON array of file paths."""
        try:
            file_list = json.loads(paths)
        except json.JSONDecodeError:
            file_list = [p.strip() for p in paths.split(",") if p.strip()]

        results = []
        for fp in file_list:
            fp_err = _check_path(fp)
            if fp_err:
                results.append(f"--- {fp} ---\nError: {fp_err}")
                continue
            p = Path(fp)
            if not p.exists():
                results.append(f"--- {fp} ---\nError: file not found")
                continue
            try:
                text = p.read_text(encoding="utf-8", errors="replace")
                results.append(f"--- {fp} ({len(text.splitlines())} lines) ---\n{text}")
            except Exception as e:
                results.append(f"--- {fp} ---\nError: {e}")
        return "\n\n".join(results)

    @server.tool()
    async def sassy_write_file(path: str, content: str, mode: str = "rewrite") -> str:
        """Write or append to a file. mode: rewrite | append"""
        err = _check_path(path)
        if err:
            return f"Error: {err}"
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        if mode == "append":
            with open(p, "a", encoding="utf-8") as f:
                f.write(content)
        else:
            p.write_text(content, encoding="utf-8")
        lines = content.count("\n") + 1
        return f"Written {lines} lines to {path} ({mode})"

    @server.tool()
    async def sassy_list_dir(path: str, depth: int = 2) -> str:
        """List directory contents. [FILE]/[DIR] prefixes. depth controls recursion."""
        err = _check_path(path)
        if err:
            return f"Error: {err}"
        depth = min(max(depth, 1), 10)
        results = []
        p = Path(path)
        if not p.is_dir():
            return f"Error: {path} is not a directory"

        def _walk(current, d, prefix=""):
            if d > depth:
                return
            try:
                entries = sorted(
                    current.iterdir(),
                    key=lambda e: (not e.is_dir(), e.name.lower()),
                )
            except PermissionError:
                results.append(f"[DENIED] {prefix}{current.name}")
                return
            limit = 100 if d > 1 else 500
            shown = 0
            total_items = 0
            for entry in entries:
                if entry.name.startswith(".") and d > 1:
                    continue
                if entry.name in ("node_modules", "__pycache__", ".git") and d > 1:
                    continue
                total_items += 1
                if shown >= limit:
                    continue
                rel = f"{prefix}{entry.name}"
                if entry.is_dir():
                    results.append(f"[DIR] {rel}")
                    _walk(entry, d + 1, f"{rel}/")
                else:
                    results.append(f"[FILE] {rel}")
                shown += 1
            if total_items > shown:
                results.append(f"  [WARNING] {total_items - shown} items hidden (showing {shown} of {total_items})")

        _walk(p, 1)
        return "\n".join(results[:1000])

    @server.tool()
    async def sassy_search_files(
        path: str,
        pattern: str,
        search_type: str = "files",
        file_pattern: str = "",
        ignore_case: bool = True,
        max_results: int = 50,
        context_lines: int = 0,
    ) -> str:
        """Search for files by name or content.

        search_type: files | content
        file_pattern: glob filter e.g. '*.py', '*.js'
        context_lines: lines of context around content matches
        """
        err = _check_path(path)
        if err:
            return f"Error: {err}"
        max_results = min(max(max_results, 1), 500)
        results = []
        p = Path(path)
        flags = re.IGNORECASE if ignore_case else 0

        if search_type == "files":
            glob_pat = file_pattern or "*"
            for match in p.rglob(glob_pat):
                if re.search(pattern, match.name, flags):
                    results.append(str(match))
                    if len(results) >= max_results:
                        break
        else:
            pat = re.compile(pattern, flags)
            glob_pat = file_pattern or "*"
            for fpath in p.rglob(glob_pat):
                if not fpath.is_file() or fpath.stat().st_size > 5_000_000:
                    continue
                try:
                    lines = fpath.read_text(encoding="utf-8", errors="replace").splitlines()
                    for i, line in enumerate(lines):
                        if pat.search(line):
                            ctx_before = lines[max(0, i - context_lines) : i]
                            ctx_after = lines[i + 1 : i + 1 + context_lines]
                            match_str = f"{fpath}:{i + 1}: {line.strip()[:200]}"
                            if context_lines > 0:
                                ctx = "\n".join(
                                    [f"  {l.strip()[:200]}" for l in ctx_before]
                                    + [f"> {line.strip()[:200]}"]
                                    + [f"  {l.strip()[:200]}" for l in ctx_after]
                                )
                                match_str = f"{fpath}:{i + 1}\n{ctx}"
                            results.append(match_str)
                            if len(results) >= max_results:
                                return "\n".join(results)
                except Exception:
                    continue

        return "\n".join(results) if results else "No matches found"

    @server.tool()
    async def sassy_move(source: str, destination: str) -> str:
        """Move or rename a file/directory."""
        for p in (source, destination):
            err = _check_path(p)
            if err:
                return f"Error: {err}"
        try:
            shutil.move(source, destination)
            return f"Moved {source} -> {destination}"
        except (OSError, shutil.Error) as e:
            return f"Error moving: {e}"

    @server.tool()
    async def sassy_copy(source: str, destination: str) -> str:
        """Copy a file or directory tree."""
        for p in (source, destination):
            err = _check_path(p)
            if err:
                return f"Error: {err}"
        try:
            src = Path(source)
            if src.is_dir():
                shutil.copytree(source, destination)
            else:
                Path(destination).parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, destination)
            return f"Copied {source} -> {destination}"
        except (OSError, shutil.Error) as e:
            return f"Error copying: {e}"

    @server.tool()
    async def sassy_file_info(path: str) -> str:
        """Get detailed file/directory metadata.

        Includes: size, timestamps, line count (text), sheet info (Excel).
        """
        err = _check_path(path)
        if err:
            return f"Error: {err}"
        p = Path(path)
        if not p.exists():
            return f"Error: {path} does not exist"
        stat = p.stat()
        info = {
            "path": str(p.resolve()),
            "type": "directory" if p.is_dir() else "file",
            "size_bytes": stat.st_size,
            "modified": stat.st_mtime,
            "created": stat.st_ctime,
        }
        if p.is_file():
            # Line count for text files
            try:
                with open(p, encoding="utf-8", errors="replace") as f:
                    line_count = sum(1 for _ in f)
                info["line_count"] = line_count
                info["last_line"] = line_count - 1
            except Exception:
                pass
            # Excel sheet metadata
            suffix = p.suffix.lower()
            if suffix in (".xlsx", ".xls", ".xlsm"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(str(p), read_only=True)
                    info["sheets"] = [
                        {"name": ws.title, "rows": ws.max_row, "cols": ws.max_column}
                        for ws in wb.worksheets
                    ]
                    wb.close()
                except Exception:
                    info["sheets"] = "(openpyxl not available)"
        elif p.is_dir():
            try:
                items = list(p.iterdir())
                info["item_count"] = len(items)
                info["files"] = sum(1 for i in items if i.is_file())
                info["directories"] = sum(1 for i in items if i.is_dir())
            except PermissionError:
                info["item_count"] = "(access denied)"

        return json.dumps(info, indent=2)

    @server.tool()
    async def sassy_mkdir(path: str) -> str:
        """Create a directory (and any missing parents)."""
        err = _check_path(path)
        if err:
            return f"Error: {err}"
        p = Path(path)
        p.mkdir(parents=True, exist_ok=True)
        return f"Created {p.resolve()}"
